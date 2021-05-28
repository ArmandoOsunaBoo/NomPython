import os

from modules import *
from openpyxl import load_workbook
from openpyxl import Workbook

class ReporteNomina:
    def __init__(self, file):
        pass
        workbook = load_workbook(filename=file, data_only=True)
        sheet = workbook["INGRESO DE DATOS"]
        workbook_movimientos = workbook["MOVIMIENTOS DE NOMINA"]
        workbook_faltas = workbook["FALTAS"]

        sheet_faltas = workbook_faltas
        sheet_movimientos = workbook_movimientos

        clave_trabajador = 0


        #tenemos que leer datos del empleado del archivo
        lunes= self.formato_fecha(sheet.cell(row=5, column=130).value)
        martes= self.formato_fecha(sheet.cell(row=5, column=131).value)
        miercoles= self.formato_fecha(sheet.cell(row=5, column=132).value)
        jueves= self.formato_fecha(sheet.cell(row=5, column=133).value)
        viernes= self.formato_fecha(sheet.cell(row=5, column=134).value)
        sabados= self.formato_fecha(sheet.cell(row=5, column=135).value)

        #El primer empleado empieza en la fila 6
        fila = 5
        #Creamos los archivos que vamos a generar para NOI y Odoo
        workbook_faltas_creado,workbook_incidencias_creado = self.crear_excels()
        # Primero llenamos el archivo de faltas con la segunda hoja del archivo subido
        sheet_faltas_creado_sheet = workbook_faltas_creado.active
        #Variables para llevar el conteo de las hojas creadas
        contador_fila_faltas_creado=1
        contador_fila_incidencias_creado = 2
        workbook_incidencias_creado_sheet = workbook_incidencias_creado.active
        mrow= sheet_faltas.max_row

        for x in range(2,mrow):
            pass
            if(sheet_faltas.cell(row=x, column=1).value!=0):
                clave_trabajador = sheet_faltas.cell(row=x, column=1).value
                #print(sheet_faltas.cell(row=x, column=4).value)
                porcentaje_l= int(round(float(sheet_faltas.cell(row=x, column=5).value)*100))
                porcentaje_ma=int(round(float(sheet_faltas.cell(row=x, column=6).value)*100))
                porcentaje_mi=int(round(float(sheet_faltas.cell(row=x, column=7).value)*100))
                porcentaje_ju=int(round(float(sheet_faltas.cell(row=x, column=8).value)*100))
                porcentaje_vi=int(round(float(sheet_faltas.cell(row=x, column=9).value)*100))
                porcentaje_sab=int(round(float(sheet_faltas.cell(row=x, column=10).value)*100))
                print("Clave Trabajador: "+str(clave_trabajador))
                print(porcentaje_l)
                print(porcentaje_ma)
                print(porcentaje_mi)
                print(porcentaje_ju)
                print(porcentaje_vi)
                print(porcentaje_sab)
                print("---------")
                print(lunes)
                print(martes)
                print(miercoles)
                print(jueves)
                print(viernes)
                print(sabados)
                print("---------------------------")




                if porcentaje_l>0:
                    pass
                    contador_fila_faltas_creado=contador_fila_faltas_creado+1
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=1).value = clave_trabajador
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=2).value = 2
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=4).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=5).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=6).value = lunes
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=7).value = lunes

                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_l
                if porcentaje_ma>0:
                    pass
                    contador_fila_faltas_creado = contador_fila_faltas_creado + 1
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_ma
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=1).value = clave_trabajador
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=2).value = 2
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=4).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=5).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=6).value = martes
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=7).value = martes

                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_ma
                if porcentaje_mi>0:
                    pass
                    contador_fila_faltas_creado = contador_fila_faltas_creado + 1
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_mi
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=1).value = clave_trabajador
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=2).value = 2
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=4).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=5).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=6).value = miercoles
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=7).value = miercoles

                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_mi
                if porcentaje_ju>0:
                    pass
                    contador_fila_faltas_creado = contador_fila_faltas_creado + 1
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_ju
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=1).value = clave_trabajador
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=2).value = 2
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=4).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=5).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=6).value = jueves
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=7).value = jueves
                if porcentaje_vi>0:
                    pass
                    contador_fila_faltas_creado = contador_fila_faltas_creado + 1
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_vi
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=1).value = clave_trabajador
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=2).value = 2
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=4).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=5).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=6).value = viernes
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=7).value = viernes
                if porcentaje_sab>0:
                    pass
                    contador_fila_faltas_creado = contador_fila_faltas_creado + 1
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=3).value = porcentaje_sab
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=1).value = clave_trabajador
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=2).value = 2
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=4).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=5).value = 0
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=6).value = sabados
                    sheet_faltas_creado_sheet.cell(row=contador_fila_faltas_creado, column=7).value = sabados

        nrow = sheet_movimientos.max_row
        print(nrow)
        contador_mov_creado=1
        for x in range(4, nrow):
            pass
            clave_trabajador = str(sheet_movimientos.cell(row=x, column=1).value)
            if clave_trabajador !="0":
                print("Clave:    "+str(clave_trabajador))
                #SUELDO P001
                print("Valores 5:"+str(sheet_movimientos.cell(row=x, column=5).value)+"||")
                if str(sheet_movimientos.cell(row=x, column=5).value) != "0":
                    pass
                    contador_mov_creado=contador_mov_creado+1
                    self.impresion_de_movimientos(contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P","1", sabados,
                                                  sheet_movimientos.cell(row=x, column=5).value)

                # SEPTIMO DÍA P109
                print("Valores 6:" + str(sheet_movimientos.cell(row=x, column=6).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=6).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "109", sabados,
                                                  sheet_movimientos.cell(row=x, column=6).value)

                #MONTO HORAS DOBLES P003
                print("Valores 7:" + str(sheet_movimientos.cell(row=x, column=7).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=7).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "3", sabados,
                                                  sheet_movimientos.cell(row=x, column=7).value)

                #MONTO HORAS TRIPLES P005
                print("Valores 8:" + str(sheet_movimientos.cell(row=x, column=8).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=8).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "5", sabados,
                                                  sheet_movimientos.cell(row=x, column=8).value)

                #VACACIONES P009
                print("Valores 9:" + str(sheet_movimientos.cell(row=x, column=9).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=9).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "9", sabados,
                                                  sheet_movimientos.cell(row=x, column=9).value)

                #PRIMA VACACIONAL P010
                print("Valores 10:" + str(sheet_movimientos.cell(row=x, column=10).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=10).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "10", sabados,
                                                  sheet_movimientos.cell(row=x, column=10).value)

                #DESCANSO LABORADO P118
                print("Valores 11:" + str(sheet_movimientos.cell(row=x, column=11).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=11).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "118", sabados,
                                                  sheet_movimientos.cell(row=x, column=11).value)

                #PRIMA DOMINICAL P019
                print("Valores 12:" + str(sheet_movimientos.cell(row=x, column=12).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=12).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "19", sabados,
                                                  sheet_movimientos.cell(row=x, column=12).value)

                #BONO POR CAPACITACION P113
                print("Valores 13:" + str(sheet_movimientos.cell(row=x, column=13).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=13).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "113", sabados,
                                                  sheet_movimientos.cell(row=x, column=13).value)

                # BONO ÁREA ESPECIAL P121
                print("Valores 14:" + str(sheet_movimientos.cell(row=x, column=14).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=14).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "121", sabados,
                                                  sheet_movimientos.cell(row=x, column=14).value)
                #BONO LIDER DE GRUPO P129
                print("Valores 15: " + str(sheet_movimientos.cell(row=x, column=15).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=15).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "129", sabados,
                                                  sheet_movimientos.cell(row=x, column=15).value)

                #BONO POR RENDIMIENTO P132
                print("Valores: 16 " + str(sheet_movimientos.cell(row=x, column=16).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=16).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "132", sabados,
                                                  sheet_movimientos.cell(row=x, column=16).value)

                #PREMIOS DE EFICIENCIA P126
                print("Valores 17:" + str(sheet_movimientos.cell(row=x, column=17).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=17).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "126", sabados,
                                                  sheet_movimientos.cell(row=x, column=17).value)

                #PREMIOS DE PUNTUALIDAD P111
                print("Valores: 18" + str(sheet_movimientos.cell(row=x, column=18).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=18).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "111", sabados,
                                                  sheet_movimientos.cell(row=x, column=18).value)

                #BONO KPI P135
                print("Valores 19:" + str(sheet_movimientos.cell(row=x, column=19).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=19).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "135", sabados,
                                                  sheet_movimientos.cell(row=x, column=19).value)

                #BONO POR ESPECIALIDAD P 133
                print("Valores 20:" + str(sheet_movimientos.cell(row=x, column=20).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=20).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "133", sabados,
                                                  sheet_movimientos.cell(row=x, column=20).value)
                #APOYO A ENCARGADO P134
                print("Valores: 21" + str(sheet_movimientos.cell(row=x, column=21).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=21).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "134", sabados,
                                                  sheet_movimientos.cell(row=x, column=21).value)

                #OTROS BONOS P114
                print("Valores 22:" + str(sheet_movimientos.cell(row=x, column=22).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=22).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "114", sabados,
                                                  sheet_movimientos.cell(row=x, column=22).value)

                #OTROS BONOS 2 P130
                print("Valores 23:" + str(sheet_movimientos.cell(row=x, column=23).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=23).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "130", sabados,
                                                  sheet_movimientos.cell(row=x, column=23).value)

                #INCENTIVO DE PRODUCTIVIDAD P131
                print("Valores 24:" + str(sheet_movimientos.cell(row=x, column=24).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=24).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "131", sabados,
                                                  sheet_movimientos.cell(row=x, column=24).value)

                #RETROACTIVO P127
                print("Valores 25:" + str(sheet_movimientos.cell(row=x, column=25).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=25).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "127", sabados,
                                                  sheet_movimientos.cell(row=x, column=25).value)

                #PERDIDA DE HERRAMIENTA D113
                print("Valores 26:" + str(sheet_movimientos.cell(row=x, column=26).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=26).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "113", sabados,
                                                  sheet_movimientos.cell(row=x, column=26).value)

                #REPOSICION GAFETE D114
                print("Valores 27:" + str(sheet_movimientos.cell(row=x, column=27).value) + "||")
                if str(sheet_movimientos.cell(row=x, column=27).value) != "0":
                    pass
                    contador_mov_creado = contador_mov_creado + 1
                    self.impresion_de_movimientos( contador_mov_creado, workbook_incidencias_creado_sheet,
                                                  clave_trabajador, "P", "114", sabados,
                                                  sheet_movimientos.cell(row=x, column=27).value)


        ruta=os.getcwd()
        print(ruta[0:3])
        try:
            os.mkdir("C:/tempx")
        except:
            pass
        workbook_incidencias_creado.save("C:/tempx/"+'out2.xls')
        os.system('start excel.exe '+'C:/tempx/out2.xls')
        workbook_faltas_creado.save("C:/tempx/"+'out.xls')
        os.system('start excel.exe '+'C:/tempx/out.xls')


        #Valores clave

        #este nadamas es para ver los valores del excel no hace nada realmente
        for columna in range (1,136):
            pass
            if columna==1:
                clave_trabajador =  sheet.cell(row=fila, column=columna).value

            valor = sheet.cell(row=fila, column=columna).value
            #print(str(columna)+" "+str(valor))

    def impresion_de_movimientos(self,fila,sheet,clave_emp,per_ded,codigo,fecha,monto):
        pass
        sheet.cell(row=fila, column=1).value = clave_emp
        sheet.cell(row=fila, column=2).value = per_ded
        sheet.cell(row=fila, column=3).value = codigo
        sheet.cell(row=fila, column=4).value = 0
        sheet.cell(row=fila, column=5).value = "N"
        sheet.cell(row=fila, column=6).value = 2
        sheet.cell(row=fila, column=7).value = fecha
        sheet.cell(row=fila, column=8).value = fecha
        sheet.cell(row=fila, column=9).value = monto


    def formato_fecha(self,fecha):
        pass
        fecha=str(fecha)
        #2021-06-05
        return fecha[8:10]+"/"+fecha[5:7]+"/"+fecha[0:4]

    def crear_excels(self):
        pass
        archivo1 = Workbook()
        sheet = archivo1.active
        sheet.cell(row=1, column=1).value = "Clave trabajador"
        sheet.cell(row=1, column=2).value = "Tipo Falta"
        sheet.cell(row=1, column=3).value = "% falta"
        sheet.cell(row=1, column=4).value = "Certificado IMSS"
        sheet.cell(row=1, column=5).value = "% pagado por IMSS"
        sheet.cell(row=1, column=6).value = "Fecha inicio"
        sheet.cell(row=1, column=7).value = "Fecha fin"
        sheet.cell(row=1, column=8).value = "Clave tipo de incidencia"
        sheet.cell(row=1, column=9).value = "Observaciones"
        archivo2 = Workbook()
        sheet = archivo2.active
        sheet.cell(row=1, column=1).value = "Clave trabajador"
        sheet.cell(row=1, column=2).value = "Per/Ded"
        sheet.cell(row=1, column=3).value = "Núm Per/Ded"
        sheet.cell(row=1, column=4).value = "Núm crédito"
        sheet.cell(row=1, column=5).value = "Aplica Destajo"
        sheet.cell(row=1, column=6).value = "Aplicación"
        sheet.cell(row=1, column=7).value = "Fecha inicio"
        sheet.cell(row=1, column=8).value = "Fecha fin"
        sheet.cell(row=1, column=9).value = "Monto o fórmula"
        sheet.cell(row=1, column=10).value = "Valor del descuento"
        sheet.cell(row=1, column=11).value = "Monto límite"
        sheet.cell(row=1, column=12).value = "Monto Acumulado"
        sheet.cell(row=1, column=13).value = "Criterio INFONAVIT"
        return archivo1,archivo2