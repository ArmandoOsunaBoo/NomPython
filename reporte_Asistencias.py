from PIL import ImageTk, Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from datetime import date
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import os

import base_de_datos

"""
Aqui se nombran la lista de las areas de empleados en las cuales deben de estar en la hoja de excel correspondiente
PRODUCTO FINAL  -  CARGADOR
GRUPO DE MANTENIMIENTO - GRUPO DE MEJORA E INSTALACION
PRODUCCION/MATERIALES
GRUPO DE CARTON
A MDF
B MDF
C MDF
D MDF
GRUPO DE EXTRUSION
GRUPO DE ENSABLE I. M.
GRUPO DE PREPARACION I.M.
GRUPO DE PREPARACION MDF
ADMINISTRACION DE PRODUCCION - ASUNTOS GENERALES - ADMINISTRACION - COMPRAS - CAPTURISTA -  CONTROL DE PRODUCCION - CHOFER - LIMPIEZA - RECURSOS HUMANOS - EMPLEADDO GENERAL - TRADUCCION - EMBARQUE/FEDEX - FINANZA - ENFERMERIA -   SISTEMAS -IMPORT / EXPORT - INVESTIGACION DE PRODUCCION - SEGURIDAD E HIGIENE
"""
thin = Side(border_style="thin", color="000000")
class ReporteAsistencias:
    book=""
    sheet=""

    def __init__(self,fecha2):
        self.fecha = fecha2
        self.texto_encabezado = self.obtener_fecha(self.fecha)
        self.book = Workbook()

        self.hojas = []

        #self.hojas.page_setup.paperSize = self.hojas.PAPERSIZE_A3

        self.hojas.append(self.book.create_sheet("PRODUCTO FINAL"))  # insert at first position
        self.hojas.append(self.book.create_sheet("GRUP. MANTENIMIENTO") )  # insert at first position
        self.hojas.append(self.book.create_sheet("PRODUCCION_MATERIALES"))   # insert at first position
        self.hojas.append( self.book.create_sheet("GRUP. CARTON"))   # insert at first position
        self.hojas.append( self.book.create_sheet("MDF A组"))   # insert at first position
        self.hojas.append( self.book.create_sheet("MDF B组"))   # insert at first position
        self.hojas.append( self.book.create_sheet("MDF C组"))   # insert at first position
        self.hojas.append( self.book.create_sheet("MDF D组"))   # insert at first position
        self.hojas.append( self.book.create_sheet("GRUP. EXTRUSION"))   # insert at first position
        self.hojas.append( self.book.create_sheet("GRUP. ENSAMBLE I.M."))   # insert at first position
        self.hojas.append( self.book.create_sheet("GRUP. PREP. I.M."))   # insert at first position
        self.hojas.append( self.book.create_sheet("GRUP. PREP. MDF"))   # insert at first position
        self.hojas.append( self.book.create_sheet("OFICINA"))   # insert at first position
        #Dar formato a las hojas del libro
        for x in range(0,13):
            pass
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(self.hojas[x],paper_size =1, orientation = 'landscape')
            self.hojas[x].sheet_properties.pageSetUpPr.fitToPage = True
            self.hojas[x].page_setup.fitToHeight = False
            self.generar_reporte(self.hojas[x],x)
        #Ingresar valores al formato dado
        for x in range(0,13):
            pass
            self.encabezados_xls(self.hojas[x])




    def insertar_datos_empleado(self,fila,nombre,numeroempleado,x):
        pass
        f=fila
        print("111 Nombre: "+nombre+" No.: "+numeroempleado)
        self.hojas[x].cell(row=f, column=2).value = nombre
        self.hojas[x].cell(row=f, column=1).value = numeroempleado
        thin = Side(border_style="thin", color="000000")

        for e in range(1,34):
            pass
            self.hojas[x].cell(row=f, column=e).border = Border(top=thin, left=thin, right=thin, bottom=thin)


    def insertar_incidencias_excel(self,fecha,fila,incidencia,x):
        pass
        fecha=str(fecha)
        dia = fecha[8:11]
        dia= int(dia)+2
        columna=dia

        print("Fila: " + str(fila) + " Columna: " + str(columna))

        self.hojas[x].cell(row=fila, column=columna).value = incidencia

        self.hojas[x].cell(row=fila, column=columna).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.hojas[x].cell(row=fila, column=columna).font = Font(color="000000")
        self.hojas[x].cell(row=fila, column=columna).fill = PatternFill(fgColor="ffffff", fill_type="solid")


    def insertar_registros_excel(self,columna,fila,nombre,fecha,evento,numeroempleado,x):
        pass
        valor_de_celda=""
        hora=fecha[11:16]
        print("FECHA- HORA: "+fecha)
        print("HORA: "+hora)
        #Tenemos que evaluar si la celda aplica para "Retardo", si llega despues de 08:00 o no tiene el registro de entrada entonces es retardo
        if evento!="Entrada" or hora>"08:00":
            valor_de_celda="R"


        #Tenemos que evaluar el día de la fecha que se manda, para poder ubicarla en el excel
        dia=fecha[8:11]
        print("La fecha es: "+fecha)
        print("El día es: "+dia)
        print("Los eventos que hay son: "+valor_de_celda)

        if valor_de_celda=="":
            valor_de_celda= "A"
        #valor_de_celda +=" - " + incidencia
        #Ya ubicado el día tenemos que registrarlo en la celda que corresponde - Hay que encontrar la columna correcta
        columna= int(columna)+1+int(dia)
        print("Fila: "+str(fila)+" Columna: "+str(columna))
        self.hojas[x].cell(row=fila, column=columna).value = valor_de_celda

        self.hojas[x].cell(row=fila, column=columna).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.hojas[x].cell(row=fila, column=columna).font = Font(color="000000")
        self.hojas[x].cell(row=fila, column=columna).fill = PatternFill(fgColor="ffffff", fill_type="solid")

    def __del__(self):
        print("se destruyo el objeto")

    def guardar_archivo(self):
        pass
        try:
            os.mkdir("C:/tempx")
        except:
            pass
        self.book.save('C:/tempx/out.xlsx')
        os.system('start excel.exe C:/tempx/out.xlsx')



    def obtener_fecha(self, fecha):
        f=str(fecha)
        f=f[5:7]
        todays_date = date.today()  # solo van a salir numeros de aqui
        todays_year = datetime.now()
        if f=="01":
            return "ENERO - "+str(todays_year.year)+"  一月份點名表"
        elif f=="02":
            return "FEBRERO - "+str(todays_year.year)+"  二月份點名表"
        elif f=="03":
            return "MARZO - "+str(todays_year.year)+"  三月份點名表"
        elif f=="04":
            return "ABRIL - "+str(todays_year.year)+"  四月份點名表"
        elif f=="05":
            return "MAYO - "+str(todays_year.year)+"  五月份點名表"
        elif f=="06":
            return "JUNIO - "+str(todays_year.year)+"  六月份點名表"
        elif f=="07":
            return "JULIO - "+str(todays_year.year)+"  七月份點名表"
        elif f=="08":
            return "AGOSTO - "+str(todays_year.year)+"  八月份點名表"
        elif f=="09":
            return "SEPTIEMBRE - "+str(todays_year.year)+"  九月份點名表"
        elif f=="10":
            return "OCTUMBRE - "+str(todays_year.year)+"  十月份點名表"
        elif f=="11":
            return "NOVIEMBRE - "+str(todays_year.year)+"  十一月份點名表"
        elif f=="12":
            return "DICIEMBRE - "+str(todays_year.year)+"  十二月份點名表"


    def generar_reporte(self,hojas,c):

        #Sección de Unir celdas
        hojas.merge_cells('A3:AG3')
        hojas.merge_cells('A4:AG4')
        #Modificaciones del documento
        width = 550
        height = 40
        ancho_fechas= 4

        cell = hojas.cell(row=3, column=1)
        if c == 0:
            subtitulo="Lista de Asistencia - PRODUCTO FINAL / CARGADORES   生管部 成品"
        elif c == 1:
            subtitulo="Lista de Asistencia - GRUPO DE MANTENIMIENTO   生技部  工務 保養組"
        elif c == 2:
            subtitulo="Lista de Asistencia - GRUPO DE PRODUCCION/MATERIALES   生管部 資材"
        elif c == 3:
            subtitulo="Lista de Asistencia - GRUPO DE CARTON    製造部 紙箱组"
        elif c == 4:
            subtitulo="Lista de Asistencia - MDF A                   MDF A组"
        elif c == 5:
            subtitulo="Lista de Asistencia - MDF B                   MDF B组"
        elif c == 6:
            subtitulo="Lista de Asistencia - MDF C                   MDF C组"
        elif c == 7:
            subtitulo="Lista de Asistencia - MDF D                   MDF D组"
        elif c == 8:
            subtitulo="Lista de Asistencia - GRUPO DE EXTRUSION    製造部 仿木 押出组"
        elif c == 9:
            subtitulo="Lista de Asistencia - GRUPO DE ENSAMBLE I.M.   製造部 仿木 組裝组"
        elif c == 10:
            subtitulo="Lista de Asistencia - GRUPO DE PREPARACION I.M.    製造部 仿木 備料组"
        elif c == 11:
            subtitulo="Lista de Asistencia - GRUPO DE PREPARACION MDF   製造部 MDF 備料组"
        elif c == 12:
            subtitulo="Lista de Asistencia - PERSONAL DE OFICINA"

        cell.value = subtitulo
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ####----------------------------------------

        cell = hojas.cell(row=4, column=1)

        cell.value = self.texto_encabezado
        cell.alignment = Alignment(horizontal='center', vertical='center')

        img = openpyxl.drawing.image.Image('./imagenes/Logo Norman.png')
        img.anchor = 'C1'
        hojas.add_image(img)
        hojas.column_dimensions['A'].width = 7
        hojas.column_dimensions['B'].width = 40

        #Generar tamaños de celdas
        if True:
            hojas.column_dimensions['C'].width = ancho_fechas
            hojas.column_dimensions['D'].width = ancho_fechas
            hojas.column_dimensions['E'].width = ancho_fechas
            hojas.column_dimensions['F'].width = ancho_fechas
            hojas.column_dimensions['G'].width = ancho_fechas
            hojas.column_dimensions['H'].width = ancho_fechas
            hojas.column_dimensions['I'].width = ancho_fechas
            hojas.column_dimensions['J'].width = ancho_fechas
            hojas.column_dimensions['K'].width = ancho_fechas
            hojas.column_dimensions['L'].width = ancho_fechas
            hojas.column_dimensions['M'].width = ancho_fechas
            hojas.column_dimensions['N'].width = ancho_fechas
            hojas.column_dimensions['O'].width = ancho_fechas
            hojas.column_dimensions['P'].width = ancho_fechas
            hojas.column_dimensions['Q'].width = ancho_fechas
            hojas.column_dimensions['R'].width = ancho_fechas
            hojas.column_dimensions['S'].width = ancho_fechas
            hojas.column_dimensions['T'].width = ancho_fechas
            hojas.column_dimensions['U'].width = ancho_fechas
            hojas.column_dimensions['V'].width = ancho_fechas
            hojas.column_dimensions['W'].width = ancho_fechas
            hojas.column_dimensions['X'].width = ancho_fechas
            hojas.column_dimensions['Y'].width = ancho_fechas
            hojas.column_dimensions['Z'].width = ancho_fechas
            hojas.column_dimensions['AA'].width = ancho_fechas
            hojas.column_dimensions['AB'].width = ancho_fechas
            hojas.column_dimensions['AC'].width = ancho_fechas
            hojas.column_dimensions['AD'].width = ancho_fechas
            hojas.column_dimensions['AE'].width = ancho_fechas
            hojas.column_dimensions['AF'].width = ancho_fechas

        #Lineas de encabezados
        hojas['A5'].value = 'NO.'
        hojas['B5'].value = 'NAME'


    def insertar_pie_xls(self,x):
            pass
            img = openpyxl.drawing.image.Image('./imagenes/reporte_nomenclatura.png')
            img.anchor = 'F40'
            self.hojas[x].add_image(img)


            #self.hojas[x]['B44'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            self.hojas[x]['B44'].border = Border(bottom=thin)
            self.hojas[x]['B45'].value = '(Nombre y Firma de supervisor)'


    def encabezados2_xls(self,sheet,fila):
        pass
        self.hojas[sheet].cell(row=fila, column=1).value = "NO."
        self.hojas[sheet].cell(row=fila, column=2).value = "NAME"
        self.hojas[sheet].cell(row=fila, column=1).font = Font(color="FFFFFF")
        self.hojas[sheet].cell(row=fila, column=1).fill = PatternFill(fgColor="00285c", fill_type="solid")
        self.hojas[sheet].cell(row=fila, column=2).font = Font(color="FFFFFF")
        self.hojas[sheet].cell(row=fila, column=2).fill = PatternFill(fgColor="00285c", fill_type="solid")
        for x in range(3,34):
            pass
            self.hojas[sheet].cell(row=fila, column=x).value = str(x-2)
            self.hojas[sheet].cell(row=fila, column=x).font = Font(color="FFFFFF")
            self.hojas[sheet].cell(row=fila, column=x).fill = PatternFill(fgColor="00285c", fill_type="solid")
            #sheet['A5'].fill = PatternFill(fgColor="00285c", fill_type="solid")


    def identificar_faltas(self,fila,sheet):
        pass
        if fila==5:
            pass
        else:
            for x in range(3, 34):
                cell = self.hojas[sheet].cell(row=fila, column=x)
                if cell.value:
                    pass
                else:
                    self.hojas[sheet].cell(row=fila, column=x).value = "F"
                    self.hojas[sheet].cell(row=fila, column=x).font = Font(color="ffffff")
                    self.hojas[sheet].cell(row=fila, column=x).fill = PatternFill(fgColor="4f4d4d", fill_type="solid")






    def encabezados_xls(self,sheet):
        #Valores de las celdas
        if True:
            sheet['C5'].value = '1'
            sheet['D5'].value = '2'
            sheet['E5'].value = '3'
            sheet['F5'].value = '4'
            sheet['G5'].value = '5'
            sheet['H5'].value = '6'
            sheet['I5'].value = '7'
            sheet['J5'].value = '8'
            sheet['K5'].value = '9'
            sheet['L5'].value = '10'
            sheet['M5'].value = '11'
            sheet['N5'].value = '12'
            sheet['O5'].value = '13'
            sheet['P5'].value = '14'
            sheet['Q5'].value = '15'
            sheet['R5'].value = '16'
            sheet['S5'].value = '17'
            sheet['T5'].value = '18'
            sheet['U5'].value = '19'
            sheet['V5'].value = '20'
            sheet['W5'].value = '21'
            sheet['X5'].value = '22'
            sheet['Y5'].value = '23'
            sheet['Z5'].value = '24'
            sheet['AA5'].value = '25'
            sheet['AB5'].value = '26'
            sheet['AC5'].value = '27'
            sheet['AD5'].value = '28'
            sheet['AE5'].value = '29'
            sheet['AF5'].value = '30'
            sheet['AG5'].value = '31'
        #Letras Color blanco
        if True:
            sheet['A5'].font = Font(color="FFFFFF")
            sheet['B5'].font = Font(color="FFFFFF")
            sheet['C5'].font = Font(color="FFFFFF")
            sheet['D5'].font = Font(color="FFFFFF")
            sheet['E5'].font = Font(color="FFFFFF")
            sheet['F5'].font = Font(color="FFFFFF")
            sheet['G5'].font = Font(color="FFFFFF")
            sheet['H5'].font = Font(color="FFFFFF")
            sheet['I5'].font = Font(color="FFFFFF")
            sheet['J5'].font = Font(color="FFFFFF")
            sheet['K5'].font = Font(color="FFFFFF")
            sheet['L5'].font = Font(color="FFFFFF")
            sheet['M5'].font = Font(color="FFFFFF")
            sheet['N5'].font = Font(color="FFFFFF")
            sheet['O5'].font = Font(color="FFFFFF")
            sheet['P5'].font = Font(color="FFFFFF")
            sheet['Q5'].font = Font(color="FFFFFF")
            sheet['R5'].font = Font(color="FFFFFF")
            sheet['S5'].font = Font(color="FFFFFF")
            sheet['T5'].font = Font(color="FFFFFF")
            sheet['U5'].font = Font(color="FFFFFF")
            sheet['V5'].font = Font(color="FFFFFF")
            sheet['W5'].font = Font(color="FFFFFF")
            sheet['X5'].font = Font(color="FFFFFF")
            sheet['Y5'].font = Font(color="FFFFFF")
            sheet['Z5'].font = Font(color="FFFFFF")
            sheet['AA5'].font = Font(color="FFFFFF")
            sheet['AB5'].font = Font(color="FFFFFF")
            sheet['AC5'].font = Font(color="FFFFFF")
            sheet['AD5'].font = Font(color="FFFFFF")
            sheet['AE5'].font = Font(color="FFFFFF")
            sheet['AF5'].font = Font(color="FFFFFF")
            sheet['AG5'].font = Font(color="FFFFFF")
        #Fondo azul
        if True:
            sheet['A5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['B5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['C5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['D5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['E5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['F5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['G5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['H5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['I5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['J5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['K5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['L5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['M5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['N5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['O5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['P5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['Q5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['R5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['S5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['T5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['U5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['V5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['W5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['X5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['Y5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['Z5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AA5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AB5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AC5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AD5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AE5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AF5'].fill = PatternFill(fgColor="00285c", fill_type="solid")
            sheet['AG5'].fill = PatternFill(fgColor="00285c", fill_type="solid")