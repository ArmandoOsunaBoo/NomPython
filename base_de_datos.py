import tkinter
import modules
import main
import os
import mysql.connector
from mysql.connector import Error
from tkinter import messagebox
import openpyxl
import pyexcel
import reporte_Asistencias
import utilidades_Generales
from tkinter.ttk import Progressbar
class DataBase:

    def __init__(self):
        pass
        try:
            connection=self.sql_connect()
        except Error as e:
            pass
        finally:
            if connection.is_connected():
                connection.close()
                print("MySQL connection is closed")

    def devolver_actualizacion(self,fecha):
        pass
        connection = mysql.connector.connect(host='192.168.108.214',
                                             port=3307,
                                             database='normanmx',
                                             user='normanmexico',
                                             password='Normanmexico1.')
        sql_select_Query = "SELECT * FROM `git_upd` WHERE ultima_actualizacion>'" + fecha + "'"
        cursor = connection.cursor()
        cursor.execute(sql_select_Query)
        records = cursor.fetchall()
        r=0
        for rowx in records:
            pass
            r=1

        return r

    def buscar_incidencia(self,fecha,numeroempleado):
        pass
        fecha=str(fecha)
        try:
            connection = mysql.connector.connect(host='192.168.108.214',
                                                 port=3307,
                                                 database='normanmx',
                                                 user='normanmexico',
                                                 password='Normanmexico1.')

            ini= fecha[0:8]+"01"
            fin= fecha[0:8]+"31"

            sql_select_Query = "SELECT * FROM `registro_incidencias` WHERE fecha>='"+ini+"' AND fecha<='"+fin+"' AND numeroempleado = '"+numeroempleado+"'"
            print(sql_select_Query)
            cursor = connection.cursor()
            cursor.execute(sql_select_Query)
            records2 = cursor.fetchall()
            c = 0
            incidenciax=""
            for rowx in records2:
                pass
                id = rowx[0]
                numero = rowx[1]
                nombre = rowx[2]
                incidencia = rowx[3]
                fecha = rowx[4]
                if incidencia!="":
                    incidenciax+=incidencia
                    print("INCIDENCIA )))))))))))))))))  "+ incidenciax)



        except Error as e:
            messagebox.showwarning("Alerta", "Hubo un error con la base de datos SQL \n" + str(e))
        finally:
            if connection.is_connected():
                connection.close()
                print("MySQL connection is closed")
                """ return incidenciax"""



    def  cargar_incidencias(self,tv):
        pass
        try:
            connection = mysql.connector.connect(host='192.168.108.214',
                                                 port=3307,
                                                 database='normanmx',
                                                 user='normanmexico',
                                                 password='Normanmexico1.')
            connection2 = mysql.connector.connect(host='192.168.108.214',
                                                  port=3307,
                                                  database='normanmx',
                                                  user='normanmexico',
                                                  password='Normanmexico1.')

            sql_select_Query2 = "SELECT * FROM `registro_incidencias`  ORDER BY fecha DESC LIMIT 100"

            cursor2 = connection2.cursor()

            cursor2.execute(sql_select_Query2)

            records2 = cursor2.fetchall()
            c=0
            for rowx in records2:
                pass
                id = rowx[0]
                numero = rowx[1]
                nombre = rowx[2]
                incidencia = rowx[3]
                fecha = rowx[4]
                # Insertar por padre
                tv.insert(parent="", index="end", iid=c, text=str(c), values=(numero, nombre,incidencia,fecha))
                c=c+1

        except Error as e:
            messagebox.showwarning("Alerta", "Hubo un error con la base de datos SQL \n" + str(e))
        finally:
            if connection2.is_connected():
                connection2.close()
                connection.close()
                print("MySQL connection is closed")

    def upload_incidences(self,filename,tv):
        pass
        cont2=0
        pyexcel.save_book_as(file_name=filename,
                             dest_file_name='C:/tempx/reciclados/datas.xlsx')
        workbook = openpyxl.load_workbook(filename='C:/tempx/reciclados/datas.xlsx')
        sheet = workbook.active
        pb = utilidades_Generales.ventana_carga()
        row_count = sheet.max_row
        conta=0
        for row in sheet.iter_rows():
            conta=conta+1
            if cont2==0:
                cont2=cont2+1
            else:
                pb.cargar_valor(((conta * 100) / row_count))
                numeromepleado = row[0].value
                numeroempleado = str(numeroempleado)
                if len(numeroempleado) == 1:
                    pass
                    numeroempleado = "0000" + numeroempleado
                elif len(numeroempleado) == 2:
                    pass
                    numeroempleado = "000" + numeroempleado
                elif len(numeroempleado) == 3:
                    pass
                    numeroempleado = "00" + numeroempleado
                elif len(numeroempleado) == 4:
                    pass
                    numeroempleado = "0" + numeroempleado

                nombre  = str(row[1].value)
                incidencia = str(row[2].value)
                fecha = str(row[3].value)
                fecha = str(fecha[0:4]+"/"+fecha[5:7]+"/"+fecha[8:10])
                print("####################################################### "+fecha)
                try:
                    connection = self.sql_connect()
                    mySql_insert_query = "INSERT INTO registro_incidencias (numeroempleado,incidencia,fecha,nombre) VALUES ('" + str(numeromepleado) + "','" + incidencia + "','" + fecha+ "','"+nombre+"') "
                    cursor = connection.cursor()
                    cursor.execute(mySql_insert_query)
                    connection.commit()
                    cursor.close()
                except mysql.connector.Error as error:
                    messagebox.showwarning("Alerta", "Error al guardar los datos 6 \n" + str(error))
                finally:
                    if connection.is_connected():
                        connection.close()
                        print("MySQL connection is closed")
                tv.delete(*tv.get_children())
                self.cargar_incidencias(tv)
        pb.cerrar_ventana()



    def upload_employees(self,filename):
        pass
        cont2=0
        conta=0
        pyexcel.save_book_as(file_name=filename,
                             dest_file_name='C:/tempx/datas2.xlsx')

        workbook = openpyxl.load_workbook(filename="C:/tempx/datas2.xlsx")
        sheet = workbook.active

        row_count = sheet.max_row
        pb = utilidades_Generales.ventana_carga()

        for row in sheet.iter_rows():
            if cont2==0:
                cont2=cont2+1
            else:
                conta=conta+1
                pb.cargar_valor( ((conta * 100)/row_count))

                numeromepleado = row[0].value
                nombre = row[1].value
                alta = row[2].value
                nacionalidad = row[3].value
                edad = row[4].value
                sexo= row[5].value
                estado_civil= row[6].value
                curp = row[7].value
                rfc = row[8].value
                municipio = row[9].value
                localidad = row[10].value
                fraccionamiento = row[11].value
                colonia = row[12].value
                calle = row[13].value
                numero_casa = row[14].value


                numeromepleado= str(numeromepleado)
                nombre = str(nombre)
                alta= str(alta)
                nacionalidad = str(nacionalidad)
                edad = str(edad)
                sexo = str(sexo)
                estado_civil = str(estado_civil)
                curp = str(curp)
                rfc = str(rfc)
                municipio = str(municipio)
                localidad = str(localidad)
                fraccionamiento = str(fraccionamiento)
                colonia = str(colonia)
                calle = str(calle)
                numero_casa = str(numero_casa)

                #val = self.sql_validar_empleado(numeromepleado)
                #print("val: "+str(val))
                if 1 == 1:
                    pass

                    self.sql_guardar_empleado(numeromepleado,nombre,alta,nacionalidad,edad,sexo,estado_civil,curp,rfc,municipio,localidad,fraccionamiento,colonia,calle,numero_casa)
                else:
                    print("jsjsj")
                    #self.sql_actualizar_empleado(numeromepleado,nombre,nacionalidad,edad,sexo,estado_civil,curp,rfc,municipio,localidad,fraccionamiento,colonia,calle,numero_casa)
        pb.cerrar_ventana()

    def sql_validar_empleado(self,numeroempleado):
        pass
        connection = self.sql_connect()
        cursor = connection.cursor()
        numeroempleado=str(numeroempleado)
        if len(numeroempleado) == 1:
            pass
            numeroempleado = "0000" + numeroempleado
        elif len(numeroempleado) == 2:
            pass
            numeroempleado = "000" + numeroempleado
        elif len(numeroempleado) == 3:
            pass
            numeroempleado = "00" + numeroempleado
        elif len(numeroempleado) == 4:
            pass
            numeroempleado = "0" + numeroempleado

        sql_select_Query = "SELECT * FROM empleados WHERE numeroempleado = '" + (numeroempleado) +"'"
        print(sql_select_Query)
        cursor.execute(sql_select_Query)
        data = cursor.fetchall()
        if len(data) == 0:
            print("Se guardan datos nuevos")
            return 0
        else:
            print('Ya hay un dato repetido')
            return 1

    def sql_guardar_empleado(self,numeroempleado,nombre,alta,nacionalidad,edad,sexo,estado_civil,curp,rfc,municipio,localidad,fraccionamiento,colonia,calle,numero_casa):
        pass
        alta = alta[8:10] + "/" + alta[5:7] + "/" + alta[0:4]
        if len(numeroempleado) == 1:
            pass
            numeroempleado = "0000" + numeroempleado
        elif len(numeroempleado) == 2:
            pass
            numeroempleado = "000" + numeroempleado
        elif len(numeroempleado) == 3:
            pass
            numeroempleado = "00" + numeroempleado
        elif len(numeroempleado) == 4:
            pass
            numeroempleado = "0" + numeroempleado
        try:
            connection = self.sql_connect()
            print(numeroempleado)
            mySql_insert_query = "INSERT INTO empleados2 (numeroempleado,nombre,alta,nacionalidad,edad,sexo,estado_civil,curp,rfc,municipio,localidad,fraccionamiento,colonia,calle,numero_casa) "+"VALUES ('" +numeroempleado+ "','" + nombre + "','"+  alta+"','"  + nacionalidad + "', '" + edad + "','"+sexo+ "','" + estado_civil+ "','" + curp+ "','" + rfc+ "','" + municipio+ "','" + localidad+"'"+ ",'" + fraccionamiento+"'"+ ",'" + colonia+""+ "','" + calle+""+ "','" + numero_casa+"' ) "

            cursor = connection.cursor()
            cursor.execute(mySql_insert_query)
            connection.commit()
            cursor.close()

        except mysql.connector.Error as error:
            messagebox.showwarning("Alerta", "---Error al guardar los datos EMPLEADOS 2:---  \n" + str(error))
        finally:
            if connection.is_connected():
                connection.close()
                print("MySQL connection is closed")

    def sql_actualizar_empleado(self,numeroempleado,nombre,nacionalidad,edad,sexo,estado_civil,curp,rfc,municipio,localidad,fraccionamiento,colonia,calle,numero_casa):
        pass
        try:
            connection = self.sql_connect()
            cursor = connection.cursor()
            if len(numeroempleado) == 1:
                pass
                numeroempleado = "0000" + numeroempleado
            elif len(numeroempleado) == 2:
                pass
                numeroempleado = "000" + numeroempleado
            elif len(numeroempleado) == 3:
                pass
                numeroempleado = "00" + numeroempleado
            elif len(numeroempleado) == 4:
                pass
                numeroempleado = "0" + numeroempleado
            # Update single record now
            sql_update_query = f"""UPDATE empleados2 set 
            nombre='{nombre}',
            nacionalidad='{nacionalidad}', 
            edad='{edad}',
            sexo='{sexo}',
            estado_civil ='{estado_civil}',
            CURP='{curp}',
            RFC='{rfc}',
            municipio ='{municipio}',
            localidad='{localidad}',
            fraccionamiento='{fraccionamiento}',
            colonia='{colonia}',
            calle='{calle}',
            numero_casa='{numero_casa}'
           
            where numero empleado = '{numeroempleado}' """

            print(sql_update_query)
            cursor.execute(sql_update_query)
            connection.commit()
            print("Record Updated successfully ")


        except mysql.connector.Error as error:
            messagebox.showwarning("Alerta", "Error al guardar los datos A \n" + str(error))
        finally:
            if connection.is_connected():
                connection.close()
                print("MySQL connection is closed")


    def upload_assistances(self,filename):
        pass
        try:
            os.mkdir("C:/tempx")
        except:
            pass
        pyexcel.save_book_as(file_name=filename,
                       dest_file_name='C:/tempx/reciclados/datas.xlsx')
        
        workbook = openpyxl.load_workbook(filename="C:/tempx/reciclados/datas.xlsx")
        sheet = workbook.active
        nombreAnterior=""
        numeromepleadoAnterior=""

        evaluacionAnterior = ""
        v_guardado=0

        cont2=0
        cont=0

        pb= utilidades_Generales.ventana_carga()

        row_count = sheet.max_row

        conta=0
        for row in sheet.iter_rows():
            if cont2==0:
                cont2=cont2+1
            else:
                conta=conta+1
                pb.cargar_valor( ((conta * 100)/row_count))

                nombre = row[0].value
                numeromepleado = row[1].value
                fecha_tiempo = row[2].value
                evento = row[3].value
                fecha_tiempo= str(fecha_tiempo)
                evaluacion=fecha_tiempo[0:11]
                print("Se guarda: "+fecha_tiempo[0:11])
                print(row[0].value)
                print(row[1].value)
                print(row[2].value)
                print(row[3].value)
                print("---------------------------------")
                print("Evaluacion: "+evaluacion)
                print("Evaluacion Anterior: "+evaluacionAnterior)
                if numeromepleadoAnterior=="":
                    pass
                    #evaluamos si no esta repetido
                    val = self.sql_validate_records(nombre, numeromepleado, fecha_tiempo, evento)
                    if val == 0:
                        # evaluamos si es entrada
                        if v_guardado == 0:
                            print("Se guardo ENTRADA...")
                            self.sql_save_records(nombre, numeromepleado, fecha_tiempo, evento)
                            v_guardado=1

                    elif val == 2:
                        messagebox.showwarning("Alerta",
                                               "Hubo un error con la operación, revise bien el archivo, si el problema persiste contacte al administrador \n")
                        break
                    else:
                        messagebox.showwarning("Alerta",
                                               "Ya existe el registro, se abortará la operación \n" + "Datos repetidos: \n" + str(
                                                   nombre + " " + numeromepleado + " " + fecha_tiempo + " " + evento))
                        break
                    numeromepleadoAnterior=numeromepleado
                elif numeromepleadoAnterior!=numeromepleado:
                    v_guardado = 0
                    val = self.sql_validate_records( nombre, numeromepleado, fecha_tiempo, evento)
                    if val == 0:
                        if v_guardado==0:
                            print("Se guardo...")
                            self.sql_save_records(nombre, numeromepleado, fecha_tiempo, evento)
                            v_guardado = 1

                    elif val == 2:
                        messagebox.showwarning("Alerta","Hubo un error con la operación, revise bien el archivo, si el problema persiste contacte al administrador \n")
                        break
                    else:
                        messagebox.showwarning("Alerta", "Ya existe el registro, se abortará la operación \n"+"Datos repetidos: \n" +str(nombre+" "+numeromepleado+" "+fecha_tiempo+" "+evento))
                        break
                    numeromepleadoAnterior = numeromepleado
        pb.cerrar_ventana()

    def sql_connect(self):
        try:
            connection = mysql.connector.connect(host='192.168.108.214',
                                             port=3307,
                                             database='normanmx',
                                             user='normanmexico',
                                             password='Normanmexico1.')
        except mysql.connector.Error as e:
            messagebox.showwarning("Alerta", "No se ha podido conectar a la base de datos SQL 3 \n"+ e)
        finally:
            pass
        return connection

    def sql_save_records(self, nombre, numeromepleado, fecha_tiempo, evento):
        cont = 0
        try:
            connection=self.sql_connect()
            mySql_insert_query = "INSERT INTO reporte_checadas (nombre, fecha_tiempo, evento,numeroempleado) VALUES ('" + nombre + "','" + fecha_tiempo + "','" + evento + "', '" + numeromepleado + "') "

            cursor = connection.cursor()
            cursor.execute(mySql_insert_query)
            connection.commit()
            print(cursor.rowcount, "Van: "+str(cont))
            cont = cont + 1
            cursor.close()

        except mysql.connector.Error as error:
            messagebox.showwarning("Alerta", "Error al guardar los datos 2 \n"+ error)
        finally:
            if connection.is_connected():
                connection.close()
                print("MySQL connection is closed")

    def sql_validate_records(self, _nombre, _numeromepleado, _fecha_tiempo, _evento):

        try:
            connection = self.sql_connect()
            cursor = connection.cursor()
            sql_select_Query = "SELECT * FROM reporte_checadas WHERE nombre = '"+(_nombre)+"' AND fecha_tiempo='"+(_fecha_tiempo) +"' AND evento= '"+(_evento)+"' AND numeroempleado = '"+( _numeromepleado)+"'"
            print(sql_select_Query)
            cursor.execute(sql_select_Query)
            data = cursor.fetchall()
            if len(data) == 0:
                print("Se guardan datos nuevos111")
                return 0
            else:
                print('Ya hay un dato repetido')
                return 1

        except:
            return 2
        finally:
            if connection.is_connected():
                connection.close()
                cursor.close()
                print("SE CIERRA LA CONEXIÓN")

    def obtener_where(self,c):
        if c == 0:
            return " area = 'PRODUCTO FINAL' OR area = 'CARGADOR' "
        elif c == 1:
            return " area = 'GRUPO DE MANTENIMIENTO' OR area = 'GRUPO DE MEJORA E INSTALACION' OR area = 'INGENIERIA'"
        elif c == 2:
            return " area = 'PRODUCCION/MATERIALES' OR area = 'MATERIALES' "
        elif c == 3:
            return " area = 'GRUPO DE CARTON' "
        elif c == 4:
            return " area = 'A MDF' "
        elif c == 5:
            return " area = 'B MDF' "
        elif c == 6:
            return " area = 'C MDF' "
        elif c == 7:
            return " area = 'D MDF' "
        elif c == 8:
            return " area = 'GRUPO DE EXTRUSION' "
        elif c == 9:
            return " area = 'GRUPO DE ENSAMBLE I.M.' "
        elif c == 10:
            return " area = 'GRUPO DE PREPARACION I.M.' "
        elif c == 11:
            return " area = 'GRUPO DE PREPARACION MDF' "
        elif c == 12:
            return """ area = 'ADMINISTRACION DE PRODUCCION' OR area = 'ASUNTOS GENERALES' OR
                       area = 'ADMINISTRACION' OR area = 'COMPRAS' OR
                       area = 'CAPTURISTA' OR area = 'CONTROL DE PRODUCCION' OR
                       area = 'CHOFER' OR area = 'LIMPIEZA' OR
                       area = 'RECURSOS HUMANOS' OR area = 'EMPLEADDO GENERAL' OR
                       area = 'TRADUCCION' OR area = 'EMBARQUE/FEDEX' OR
                       area = 'FINANZA' OR area = 'DEPARTAMENTO MEDICO' OR
                       area = 'SISTEMAS' OR area = 'IMPO / EXPO' OR
                       area = 'INVESTIGACION DE PRODUCCION' OR area = 'SEGURIDAD E HIGIENE' OR
                       area = 'PROGRAMACION DIARIA' """

    def eliminado_incidencias(self,date1,date2,tv):
        pass
        try:
            connection = mysql.connector.connect(host='192.168.108.214',
                                                 port=3307,
                                                 database='normanmx',
                                                 user='normanmexico',
                                                 password='Normanmexico1.')
            sql_delete_Query = "DELETE FROM `registro_incidencias` WHERE `fecha`>='" + date1 + "' AND `fecha`<='"+date2+"'"
            cursor = connection.cursor()
            cursor.execute(sql_delete_Query)
            connection.commit()
            print('number of rows deleted', cursor.rowcount)
            tv.delete(*tv.get_children())
            self.cargar_incidencias(tv)
        except mysql.connector.Error as error:
            print("Failed to delete record from table: {}".format(error))
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
                print("MySQL connection is closed")



    def reporte_asistencias(self,fecha1,fecha2):
        archivoxls = reporte_Asistencias.ReporteAsistencias(fecha2)
        try:
            connection = mysql.connector.connect(host='192.168.108.214',
                                                 port=3307,
                                                 database='normanmx',
                                                 user='normanmexico',
                                                 password='Normanmexico1.')
            connection2 = mysql.connector.connect(host='192.168.108.214',
                                                 port=3307,
                                                 database='normanmx',
                                                 user='normanmexico',
                                                 password='Normanmexico1.')
            connection3 = mysql.connector.connect(host='192.168.108.214',
                                                  port=3307,
                                                  database='normanmx',
                                                  user='normanmexico',
                                                  password='Normanmexico1.')


            #SELECT * FROM `reporte_checadas` WHERE `fecha_tiempo`>='2021/04/30 00:00:00' AND `fecha_tiempo`<='2021/04/30 23:59:59 ORDER BY nombre,fecha_tiempo ASC'

            nombre=""
            nombreAnterior=""
            fecha=""
            fecha_anterior=""
            evento=""
            eventoAnterior=""
            numeroempleado=""
            numeroempleadoAnterior=""
            contador_empleados=0
            columna=1
            fila=5
            flujo=0
            fila_del_empleado=0
            #Aqui inicia el conteo de hojas de excel por áreas de la empresa
            bp = utilidades_Generales.ventana_carga()
            for x in range(0,13):
                pass
                contador_empleados = 0
                columna = 1
                fila = 5
                sentencia_WHERE=self.obtener_where(x)

                bp.cargar_valor(((x*100)/13))

                sql_select_Query2 = "SELECT nombre,numeroempleado FROM `empleados` WHERE "+sentencia_WHERE+" AND `checador` IS NOT NULL ORDER BY nombre ASC"
                print("XXX---"+sql_select_Query2)
                cursor2 = connection2.cursor()
                # Cursor 1 son checadas y cursor 2 son lista de empleados

                cursor2.execute(sql_select_Query2)
                # get all records
                archivoxls.insertar_pie_xls(x)
                records2 = cursor2.fetchall()
                #Conteo de empleados
                for rowx in records2:
                    pass
                    flujo=0
                    contador_empleados = contador_empleados+1
                    fila = fila + 1
                    fila_del_empleado= fila
                    name= rowx[0]
                    numero = rowx[1]
                    #Aqui tenemos que iterar
                    sql_select_Query = "SELECT * FROM `reporte_checadas` WHERE `numeroempleado`='"+numero+"' AND `fecha_tiempo`>='" + fecha1 + " 00:00:00' AND `fecha_tiempo`<='" + fecha2 + " 23:59:59' ORDER BY nombre,fecha_tiempo ASC"
                    print(sql_select_Query)
                    cursor = connection.cursor()
                    cursor.execute(sql_select_Query)
                    records = cursor.fetchall()
                    #Llenamos los datos del empleado
                    print("4444 Nombre: " + name + " No.: " + numeroempleado)
                    if contador_empleados==38:
                        pass
                        flujo=1
                        fila = fila + 9
                        archivoxls.encabezados2_xls(x,fila)
                        fila = fila + 1
                        fila_del_empleado=fila
                        columna=1
                        contador_empleados = 0
                    archivoxls.insertar_datos_empleado(fila,name,numero,x)
                    #Aqui tenemos que iterar por cada día que pase de la selección
                    #Este siempre va ser 1 reccorido solamente en este for
                    row_count = cursor.rowcount
                    print("--------------Numero de filas encontradas: {}".format(row_count))
                    if row_count == 0:
                        print("No hay Registros de este empleado")
                        fila= fila -1
                        flujo=-1
                    else:
                        #Conteo de checadas del día
                        for row in records:
                            nombre= row[1]
                            fecha=row[2]
                            fecha=str(fecha)
                            #fecha=fecha[0:11]
                            evento= row[3]
                            numeroempleado= row[4]
                            #Evaluamos si cambio de numero de empleado, si es el mismo o si es el primer empledo a evaluar

                            #Aqui como la fecha es diferente hay que imprimir en el excel, después se busca una incidencia
                            print("Fila: "+str(fila))
                            print("Columna: " + str(columna))
                            print("2222 Nombre: " + nombre + " No.: " + numeroempleado)
                            archivoxls.identificar_faltas(fila, x)
                            archivoxls.insertar_registros_excel(columna,fila,nombre,fecha,evento,numeroempleado,x)
                    #Aqui continuamos con los datos del empleado pero ahora se calculan las incidencias
                    #self.buscar_incidencia(fecha1, numeroempleado)
                    connection3 = mysql.connector.connect(host='192.168.108.214',
                                                         port=3307,
                                                         database='normanmx',
                                                         user='normanmexico',
                                                         password='Normanmexico1.')
                    ini = fecha1[0:8] + "01"
                    fin = fecha1[0:8] + "31"
                    print("@")
                    sql_select_Query3 = "SELECT * FROM `registro_incidencias` WHERE fecha>='" + ini + "' AND fecha<='" + fin + "' AND numeroempleado = '" + numeroempleado + "'"
                    print(sql_select_Query3)
                    cursor3 = connection3.cursor()
                    cursor3.execute(sql_select_Query3)
                    records3 = cursor3.fetchall()
                    #Incidencias del empleado
                    for row3 in records3:
                        pass
                        id = row3[0]
                        numero = row3[1]
                        nombre = row3[2]
                        incidencia = row3[3]
                        fecha = row3[4]
                        print("SE ENCONTRO INCIDENCIA______ "+nombre+" "+numeroempleado+" "+incidencia)
                        if incidencia != "":
                            archivoxls.insertar_incidencias_excel(fecha,fila,incidencia,x)


            bp.cerrar_ventana()

        except Error as e:
            messagebox.showwarning("Alerta", "Hubo un error con la base de datos SQL \n"+ str(e))
        finally:
            if connection2.is_connected():
                connection2.close()
                connection.close()
                print("MySQL connection is closed")

        archivoxls.guardar_archivo()


    def borrado_asistencias(self,fecha1,fecha2):
        archivoxls = reporte_Asistencias.ReporteAsistencias(fecha2)
        try:
            connection = mysql.connector.connect(host='192.168.108.214',
                                                 port=3307,
                                                 database='normanmx',
                                                 user='normanmexico',
                                                 password='Normanmexico1.')


            sql_Delete_query = "DELETE FROM `reporte_checadas` WHERE `fecha_tiempo`>='"+fecha1+"' AND `fecha_tiempo`<='"+fecha2+"'"
            cursor = connection.cursor()
            cursor.execute(sql_Delete_query)
            connection.commit()

        except Error as e:
            messagebox.showwarning("Alerta", "Hubo un error con la base de datos SQL \n"+ str(e))
        finally:
            if connection.is_connected():
                connection.close()
                print("MySQL connection is closed")

